# SPDX-License-Identifier: Apache-2.0
# Copyright (C) 2023-2026 Sebastien Rousseau. All rights reserved.

"""Tests for the pain001-loader-xlsx plugin.

The plugin is validated against the *external* pain001 plugin
contract (:mod:`pain001.plugins`) - exactly the same surface any
other third-party loader uses. A regression in the contract is
caught here before users hit it.
"""

from __future__ import annotations

import openpyxl
import pytest
from pain001.plugins import (
    PAIN001_API_VERSION,
    AbstractLoader,
    LoaderResult,
    PluginMeta,
)

from pain001_loader_xlsx import XlsxLoader, __version__


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------
@pytest.fixture
def xlsx_file(tmp_path):
    """A two-row payment .xlsx with text-typed IBANs and a header row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "amount", "debtor_account_IBAN"])
    ws.append(["MSG-001", 100.00, "DE89370400440532013000"])
    ws.append(["MSG-002", 250.50, "FR1420041010050500013M02606"])
    path = tmp_path / "payments.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def xlsx_with_numeric_iban(tmp_path):
    """An .xlsx whose debtor_account_IBAN column is numeric (the bad case)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "amount", "debtor_account_IBAN"])
    ws.append(["MSG-001", 100.00, 89370400440532013000])  # numeric IBAN
    path = tmp_path / "broken-payments.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def xlsx_empty_sheet(tmp_path):
    """An .xlsx whose first sheet is completely empty."""
    wb = openpyxl.Workbook()
    # The auto-created sheet is empty; just save.
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    return str(path)


# ---------------------------------------------------------------------------
# Contract surface
# ---------------------------------------------------------------------------
def test_xlsx_loader_satisfies_abstract_loader_protocol():
    """XlsxLoader is structurally an AbstractLoader (no subclass needed)."""
    assert isinstance(XlsxLoader(), AbstractLoader)


def test_meta_advertises_correct_name_extensions_and_api_version():
    """meta + extensions match the registry's lookup expectations."""
    loader = XlsxLoader()
    assert isinstance(loader.meta, PluginMeta)
    assert loader.meta.name == "xlsx"
    assert loader.meta.api_version == PAIN001_API_VERSION
    assert loader.extensions == (".xlsx", ".xlsm")
    # `source` is auto-stamped by pain001 at entry-point registration;
    # for a direct (non-discovered) construction it stays "built-in"
    # as the default - this is fine because the test instantiates the
    # plugin directly rather than through the registry.
    assert loader.meta.source in {
        "built-in",
        f"pain001-loader-xlsx=={__version__}",
    }


# ---------------------------------------------------------------------------
# load(): happy path + edge cases
# ---------------------------------------------------------------------------
def test_load_returns_one_dict_per_data_row(xlsx_file):
    """A header row + N data rows -> N dicts."""
    result = XlsxLoader().load(xlsx_file)
    assert isinstance(result, LoaderResult)
    assert result.source_hint == xlsx_file
    assert len(result.rows) == 2
    assert result.rows[0]["id"] == "MSG-001"
    assert result.rows[0]["debtor_account_IBAN"] == "DE89370400440532013000"
    assert result.rows[1]["amount"] == 250.50


def test_load_refuses_numeric_iban_cell(xlsx_with_numeric_iban):
    """An Excel "General" IBAN cell raises with a clear remediation."""
    with pytest.raises(ValueError, match="Excel's 'General' cell format"):
        XlsxLoader().load(xlsx_with_numeric_iban)


def test_load_raises_on_empty_sheet(xlsx_empty_sheet):
    """A workbook whose first sheet has no header row is rejected."""
    with pytest.raises(ValueError, match="no header row"):
        XlsxLoader().load(xlsx_empty_sheet)


def test_load_raises_on_workbook_with_no_sheets(tmp_path, monkeypatch):
    """Defensive: a workbook with zero sheets surfaces a clear error."""

    class _StubWorkbook:
        sheetnames = []

        def __getitem__(self, name):
            raise KeyError(
                name
            )  # pragma: no cover - guarded by sheetnames check

        def close(self):
            pass

    monkeypatch.setattr(
        "pain001_loader_xlsx.loader.openpyxl.load_workbook",
        lambda *a, **kw: _StubWorkbook(),
    )
    with pytest.raises(ValueError, match="no sheets"):
        XlsxLoader().load(str(tmp_path / "fake.xlsx"))


def test_load_handles_header_row_with_blank_columns(tmp_path):
    """Cells with ``None`` headers are preserved under the empty-string key."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", None, "amount"])
    ws.append(["MSG-001", "ignored", 100.00])
    path = tmp_path / "blanks.xlsx"
    wb.save(path)
    result = XlsxLoader().load(str(path))
    # The empty header becomes "" in our dict.
    assert "" in result.rows[0]
    assert result.rows[0]["id"] == "MSG-001"
    assert result.rows[0]["amount"] == 100.00


# ---------------------------------------------------------------------------
# load_streaming(): chunking
# ---------------------------------------------------------------------------
def test_load_streaming_yields_chunks_of_at_most_chunk_size(xlsx_file):
    """chunk_size=1 against a 2-row sheet yields two LoaderResult chunks."""
    chunks = list(XlsxLoader().load_streaming(xlsx_file, chunk_size=1))
    assert len(chunks) == 2
    assert all(isinstance(c, LoaderResult) for c in chunks)
    assert all(len(c.rows) == 1 for c in chunks)


def test_load_streaming_final_chunk_may_be_smaller(xlsx_file):
    """chunk_size=3 against a 2-row sheet yields one chunk of size 2."""
    chunks = list(XlsxLoader().load_streaming(xlsx_file, chunk_size=3))
    assert len(chunks) == 1
    assert len(chunks[0].rows) == 2


def test_load_streaming_source_hint_preserved(xlsx_file):
    """Every emitted chunk carries the original source path."""
    chunks = list(XlsxLoader().load_streaming(xlsx_file, chunk_size=1))
    assert {c.source_hint for c in chunks} == {xlsx_file}


def test_load_streaming_propagates_iban_guard(xlsx_with_numeric_iban):
    """The streaming variant also enforces the IBAN-as-text rule."""
    with pytest.raises(ValueError, match="Excel's 'General' cell format"):
        list(XlsxLoader().load_streaming(xlsx_with_numeric_iban, chunk_size=10))


# ---------------------------------------------------------------------------
# Registry integration: pain001 actually dispatches to this loader
# ---------------------------------------------------------------------------
def test_registry_dispatches_xlsx_to_this_plugin():
    """pain001's registry, populated lazily, finds this loader by extension."""
    from pain001.plugins.registry import registry

    registry.reset()
    # Force population (which runs entry-point discovery).
    registry._ensure_populated()
    loader = registry.get_loader_for_extension(".xlsx")
    assert loader is not None
    assert loader.meta.name == "xlsx"
    # When discovered via entry point, source is the dist string.
    assert loader.meta.source.startswith("pain001-loader-xlsx==")
