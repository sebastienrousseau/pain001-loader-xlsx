# SPDX-License-Identifier: Apache-2.0
# Copyright (C) 2023-2026 Sebastien Rousseau. All rights reserved.

"""Excel (.xlsx / .xlsm) loader plugin for pain001.

Reads payment rows from the first sheet of an OOXML spreadsheet,
treating row 1 as the header and rows 2..N as records. Cells are
read with ``data_only=True`` so formulas evaluate to their cached
last-saved value (the same value the user sees in Excel).

Implements the :class:`pain001.plugins.AbstractLoader` Protocol
without subclassing - structural typing means this plugin only
needs the four declared attributes (``meta``, ``extensions``,
``load``, ``load_streaming``) to satisfy the contract.

The loader is registered with pain001 via the ``pain001.loaders``
entry-point group in :file:`pyproject.toml`; users get .xlsx
dispatch automatically after ``pip install pain001-loader-xlsx``.

IBAN handling
-------------
Excel's General cell format silently strips leading zeros from
numeric-looking strings - a known data-corruption mode in SAP /
Oracle / Workday exports. To protect against this the loader
refuses any row whose ``debtor_account_IBAN`` /
``creditor_account_IBAN`` / ``charge_account_IBAN`` cell is typed
as a number (Excel ``General``) and asks the user to re-type the
column as ``Text`` first. Caught early, the warning saves the user
from wiring an IBAN with a missing digit to a bank.
"""

from __future__ import annotations

from collections.abc import Iterable
from importlib import metadata
from typing import Any

import openpyxl
from pain001.plugins import (
    PAIN001_API_VERSION,
    LoaderResult,
    PluginMeta,
)

# Columns whose values are bank identifiers and must never be read
# from an Excel "General" cell (Excel strips leading zeros from
# numeric-looking strings). Lowercased for case-insensitive matching
# against the header row.
_IBAN_HEADERS = frozenset(
    {
        "debtor_account_iban",
        "creditor_account_iban",
        "charge_account_iban",
    }
)


def _plugin_version() -> str:
    """Return the package version (or ``"0.0.0"`` for editable dev installs).

    Returns:
        SemVer string from importlib.metadata when the package is
        installed; ``"0.0.0"`` when running from a checkout that hasn't
        been pip-installed yet.
    """
    try:
        return metadata.version("pain001-loader-xlsx")
    except metadata.PackageNotFoundError:  # pragma: no cover - dev install
        return "0.0.0"


class XlsxLoader:
    """Read flat-record payment data from an Excel ``.xlsx`` / ``.xlsm`` file.

    Satisfies the structural :class:`pain001.plugins.AbstractLoader`
    Protocol. Construction takes no arguments; the registry
    instantiates it once per process at first lookup.

    Sheet selection: the first sheet wins. Cross-sheet payment
    batches are out of scope; users with that requirement should
    consolidate to a single sheet (or open an issue describing the
    layout).
    """

    meta = PluginMeta(
        name="xlsx",
        version=_plugin_version(),
        description=(
            "Read flat-record payment data from Excel .xlsx / .xlsm files."
        ),
        api_version=PAIN001_API_VERSION,
    )
    extensions = (".xlsx", ".xlsm")

    def load(self, path: str) -> LoaderResult:
        """Read every data row from ``path`` and return them all at once.

        Args:
            path: Filesystem path to the spreadsheet.

        Returns:
            A :class:`LoaderResult` carrying one dict per data row
            (header-cell-value -> data-cell-value) and the original
            path as source hint.

        Raises:
            ValueError: When the workbook has no sheets, no header
                row, or when an IBAN column has been typed as a
                numeric cell.
        """
        rows = list(self._iter_rows(path))
        return LoaderResult(rows=rows, source_hint=path)

    def load_streaming(
        self, path: str, chunk_size: int
    ) -> Iterable[LoaderResult]:
        """Yield ``chunk_size``-row :class:`LoaderResult` chunks.

        Args:
            path: Filesystem path to the spreadsheet.
            chunk_size: Maximum rows per yielded result.

        Returns:
            An iterable of :class:`LoaderResult`; the final chunk may
            be smaller than ``chunk_size``.
        """
        buf: list[dict[str, Any]] = []
        for record in self._iter_rows(path):
            buf.append(record)
            if len(buf) >= chunk_size:
                yield LoaderResult(rows=buf, source_hint=path)
                buf = []
        if buf:
            yield LoaderResult(rows=buf, source_hint=path)

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------
    def _iter_rows(self, path: str) -> Iterable[dict[str, Any]]:
        """Stream rows from the first sheet, refusing IBAN-as-number cells.

        Args:
            path: Filesystem path to the spreadsheet.

        Yields:
            One dict per data row, keyed by the header-row values.

        Raises:
            ValueError: When the workbook has no sheets, no header
                row, or when an IBAN column is numeric.
        """
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if not workbook.sheetnames:
            raise ValueError(f"workbook {path!r} contains no sheets")
        worksheet = workbook[workbook.sheetnames[0]]
        rows_iter = worksheet.iter_rows()
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise ValueError(
                f"workbook {path!r} sheet 1 is empty (no header row)"
            ) from exc
        headers = [
            (cell.value if cell.value is not None else "")
            for cell in header_row
        ]
        iban_columns = {
            idx
            for idx, name in enumerate(headers)
            if str(name).strip().lower() in _IBAN_HEADERS
        }
        for row in rows_iter:
            self._guard_iban_cells(row, iban_columns, headers, path)
            yield dict(
                zip(
                    headers,
                    (cell.value for cell in row),
                    strict=False,
                )
            )
        workbook.close()

    def _guard_iban_cells(
        self,
        row: tuple[Any, ...],
        iban_columns: set[int],
        headers: list[Any],
        path: str,
    ) -> None:
        """Raise when an IBAN column cell carries a numeric type.

        Args:
            row: The current data row's cells (openpyxl ``Cell``
                instances in read-only mode).
            iban_columns: Indices of header columns whose values are
                IBANs.
            headers: The header-row values (for the error message).
            path: The source file path (for the error message).

        Raises:
            ValueError: When any IBAN column cell is numeric.
        """
        for idx in iban_columns:
            if idx >= len(row):  # pragma: no cover - openpyxl pads rows
                # Defensive: openpyxl's read_only mode auto-pads rows
                # to the header width in practice, but the docs don't
                # guarantee it. The guard protects against a hand-
                # rolled iterator that strips trailing empty cells.
                continue
            value = row[idx].value
            if isinstance(value, int | float):
                column_name = headers[idx]
                raise ValueError(
                    f"workbook {path!r} column {column_name!r} contains a "
                    f"numeric value ({value!r}) where an IBAN string is "
                    "expected. Excel's 'General' cell format silently "
                    "strips leading zeros from IBANs; re-type the column "
                    "as 'Text' (in Excel: select the column, "
                    "Format Cells -> Number -> Text) and re-export."
                )
